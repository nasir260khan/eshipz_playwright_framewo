import random
import string
from pathlib import Path
from openpyxl import load_workbook
from playwright.sync_api import expect
from utils.config import BASE_URL

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "downloads" / "warehouse_template.xlsx"

WAREHOUSE_TYPES = [
    ("Pickup", "pickup"),
    ("Receiver", "delivery"),
    ("RTO", "rto"),
]


def phone():
    return "9" + "".join(str(random.randint(0, 9)) for _ in range(9))


def random_alias(display_type):
    code = "".join(random.choices(string.ascii_uppercase, k=6))

    if display_type == "Pickup":
        return f"PK{code}"
    elif display_type == "Receiver":
        return f"RC{code}"
    else:
        return f"RT{code}"


def open_warehouse_list(page):
    page.goto(
        f"{BASE_URL}/v2/settings/shipping-setup/warehouse-list",
        wait_until="domcontentloaded"
    )
    page.wait_for_timeout(3000)
    expect(page.get_by_role("button", name="Bulk Upload")).to_be_visible(timeout=30000)


def warehouse_data(display_type, warehouse_type):
    alias = random_alias(display_type)
    email_num = random.randint(10000, 99999)

    return {
        "Display Type": display_type,
        "Warehouse Type": warehouse_type,
        "Warehouse ID": "",
        "Company Name": f"{display_type} Excel Company",
        "Contact Name": f"{display_type} Excel Contact",
        "Email": f"{display_type.lower()}{email_num}@gmail.com",
        "Phone": phone(),
        "Street 1": f"{display_type} Excel Street 1",
        "Street 2": f"{display_type} Excel Street 2",
        "Street 3": f"{display_type} Excel Street 3",
        "City": "Delhi",
        "State": "Delhi",
        "Country": "India",
        "Pincode": "110020",
        "Address Type": "Residential",
        "Tax ID/GSTIN": "",
        "Fax": "123456",
        "Alias Name": alias,
        "what3words": "test word auto",
    }


def create_filled_excel_from_template():
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]

    rows = [
        warehouse_data(display, value)
        for display, value in WAREHOUSE_TYPES
    ]

    for row_index, data in enumerate(rows, start=2):
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=row_index, column=col_index).value = data.get(header, "")

    output_path = Path("downloads") / f"filled_warehouse_template_{random.randint(1000, 9999)}.xlsx"
    wb.save(output_path)

    print(f"✅ Excel file created: {output_path}")
    print("✅ Alias names are below 20 characters")

    return output_path, rows


def upload_excel(page, excel_path):
    open_warehouse_list(page)

    page.get_by_role("button", name="Bulk Upload").click()
    page.wait_for_timeout(1000)

    page.locator("input[type='file']").set_input_files(str(excel_path.resolve()))
    page.wait_for_timeout(2000)

    page.get_by_role("button", name="Upload", exact=True).click()
    page.wait_for_timeout(5000)

    print(f"✅ Excel uploaded: {excel_path}")


def search_by_alias(page, alias):
    open_warehouse_list(page)

    search = page.get_by_role("textbox", name="Search by name, address...")
    search.click()
    search.fill("")
    search.fill(alias)

    page.wait_for_timeout(3000)
    expect(page.get_by_text(alias, exact=False).first).to_be_visible(timeout=15000)

    print(f"✅ Warehouse found by alias: {alias}")


def validate_ui(page, data):
    alias = data["Alias Name"]

    search_by_alias(page, alias)

    page.get_by_role("button", name="Edit Warehouse").first.click()
    page.wait_for_timeout(2000)

    expect(page.get_by_role("textbox", name="Alias Name *")).to_have_value(alias, timeout=10000)
    expect(page.get_by_role("textbox", name="Company Name *")).to_have_value(data["Company Name"], timeout=10000)
    expect(page.get_by_role("textbox", name="Contact Name *")).to_have_value(data["Contact Name"], timeout=10000)
    expect(page.get_by_role("textbox", name="Email Address *")).to_have_value(data["Email"], timeout=10000)
    expect(page.get_by_role("textbox", name="Phone Number *")).to_have_value(data["Phone"], timeout=10000)
    expect(page.get_by_role("textbox", name="Address Line 1 *")).to_have_value(data["Street 1"], timeout=10000)
    expect(page.get_by_role("textbox", name="City *")).to_have_value(data["City"], timeout=10000)
    expect(page.get_by_role("textbox", name="Pincode *")).to_have_value(data["Pincode"], timeout=10000)

    print(f"✅ {data['Display Type']} validated on UI: {alias}")

    open_warehouse_list(page)


def upload_and_validate_warehouse_excel(page):
    excel_path, rows = create_filled_excel_from_template()

    upload_excel(page, excel_path)

    for row in rows:
        validate_ui(page, row)

    return rows