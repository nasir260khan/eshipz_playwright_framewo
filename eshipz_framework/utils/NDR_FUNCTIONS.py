# ----------------get shipmet details like customer reference-------------------------
def get_created_shipment_details(page):
    custref = page.locator("#customer-reference").input_value()
    print("Customer Ref:", custref)

    awb_text = page.locator("p:has-text('AWB:')").inner_text().split(":")[-1].strip()
    print("AWB:", awb_text)

    page.get_by_role("button", name="Close").click()
    return custref, awb_text


# ----------------------serac the orders in shipment through bulk --------------------------
def search_shipment_by_customer_ref(page, base_url, custref):
    page.goto(f"{base_url}/v2/fulfillment/shipment/")
    page.wait_for_load_state("networkidle")

    page.get_by_role("button", name="Bulk Search").click()
    page.get_by_role("textbox", name="Search").fill(custref)
    page.get_by_role("button", name="Search shipments").click()

def search_shipment_by_customer_refexcep(page, base_url, custref,awb_text):
    page.goto(f"{base_url}/v2/exceptions/ndr/actionRequired")
    page.wait_for_load_state("networkidle")

    page.get_by_role("button", name="Bulk Search").click()
    page.get_by_role("textbox", name="Search").fill(custref)
    page.get_by_role("button", name="Search exception").click()
    page.wait_for_timeout(3000)

    # check if customer ref exists in table
    shipment_locator = page.locator("td").filter(has_text=str(awb_text))

    print("Searching AWB:", awb_text)
    print("Matched Count:", shipment_locator.count())

    if shipment_locator.count() > 0:
        print(f"Shipment present in NPR Action Required page with shipment AWB number: {awb_text}")
    else:
        print(f"Shipment NOT found in NPR Action Required page with shipment AWB number: {awb_text}")


def update_tracking_exception(page):
    print("➡ Updating tracking status to Exception")

    page.get_by_text("Info Received").first.click()

    page.get_by_role("button", name="Tracking", exact=True).click()

    page.get_by_role("button", name="Update Tracking").click()

    page.get_by_role("combobox", name="Tracking Status *").click()
    page.get_by_role("option", name="Exception").click()

    page.get_by_role("combobox", name="Tracking Sub Status *").click()
    page.get_by_role("option", name="Undelivered").click()

    page.get_by_role("button", name="Update", exact=True).click()

    print("✅ Tracking updated successfully")


# -----------------------performing npractions-----------------------------------
def open_ndr_action(page):
    page.locator("td > div > div:nth-child(2)").click()
    page.get_by_role("button", name="NDR", exact=True).click()
    page.locator("#ndr-action").click()


# ---------------- reshcheduleing action for npr-----------------------------

from datetime import datetime
import random

def fill_NDR_reschedule_date(page):

    today = datetime.now()

    # only between 10 and 11
    hour = random.choice([10, 11])
    minute = random.randint(0, 59)

    dt = today.replace(hour=hour, minute=minute, second=0)

    # IMPORTANT FORMAT
    formatted_date = dt.strftime("%Y-%m-%dT%H:%M")

    print("Selected Time:", formatted_date)

    date_input = page.locator("#reattempt-date")

    # directly set value
    date_input.fill(formatted_date)

    # trigger events
    date_input.dispatch_event("input")
    date_input.dispatch_event("change")

    page.locator("textarea").click()

    # page.get_by_role("button", name="Confirm").click()


def fill_NDR_edit_and_reschedule(page):
    page.get_by_role("button", name="Change Details").click()
    page.get_by_role("textbox", name="Customer Name").click()
    page.get_by_role("textbox", name="Customer Name").fill("hhw")
    page.get_by_placeholder("Phone Number").click()
    page.get_by_placeholder("Phone Number").fill("78787878787")
    page.get_by_role("textbox", name="Customer Address").click()
    page.get_by_role("textbox", name="Customer Address").fill("hvbwkfjwenf")
    page.get_by_placeholder("Pincode").click()
    page.get_by_placeholder("Pincode").fill("110002")




# --------CANCEL ACTION POP UP AND VALIDATIG--------------------------------------
def RTOFUNC(page):
    page.get_by_role("button", name="RTO").nth(1).click()
    page.get_by_role("button", name="Confirm").click()

    success_msg = page.get_by_text(
        "NDR processed successfully"
    ).last

    expect(success_msg).to_be_visible(timeout=10000)

    popup_text = success_msg.inner_text().strip()
    print("Cancel Popup Message:", popup_text)

    assert "NDR processed successfully" in popup_text

    print("✅ RTO  success validated")


# ------------------------------------------------------------------------------

import time
from playwright.sync_api import TimeoutError

import time
from playwright.sync_api import TimeoutError


# -------------------ACESSEING npraction  POP UP AND VALIDATING THE POP UP INNERTEXT------
# ----------and aslo if the shipment is failed then it willbe under action required tab---
def confirm_ndr_action(page, awb_text=None, action_type=None):
    if action_type in ["RE-ATTEMPT", "EDIT_AND_RESCHEDULE"]:
        page.get_by_role("button", name="Confirm").click()
        time.sleep(3)

    success_msg = page.locator(
        "span:has-text('NDR Instruction has been successfully shared to the courier partner')"
    ).first

    try:
        success_msg.wait_for(state="visible", timeout=5000)

        popup_text = success_msg.inner_text().strip()
        print("Popup Message:", popup_text)

        # ASSERT
        assert "NDR Instruction has been successfully shared to the courier partner" in popup_text

        print("✅ NDR action successful")
        print("✅ NDR ACTION Success")

    except TimeoutError:
        print("⚠️ sucessmessage not displayed → going to Action Required page")

        page.goto("https://uat.eshipz.com:444/v2/exceptions/ndr/actionRequired")
        page.wait_for_load_state("networkidle")

        # ✅ perform search using AWB
        if awb_text:
            page.get_by_role("button", name="Bulk Search").click()

            search_box = page.get_by_role("textbox", name="Search")
            search_box.wait_for(state="visible")

            search_box.fill(awb_text)

            page.get_by_role("button", name="Search exception").click()

            print(f"🔍 Searched AWB in Action Required: {awb_text}")
            return "ACTION_REQUIRED"

    time.sleep(3)


# ----------search shipment through bulk npr uder action requested tab ------------------
def search_awb_in_ndr_action_requested(page, base_url, awb_text):
    page.goto(f"{base_url}/v2/exceptions/ndr/actionRequested")
    page.wait_for_load_state("networkidle")

    page.get_by_role("button", name="Bulk Search").click()
    page.get_by_role("textbox", name="Search").fill(awb_text)
    page.get_by_role("button", name="Search exception").click()

    time.sleep(4)

    row = page.locator(
        f"//div[contains(@class,'flex') and .//text()[contains(.,'{awb_text}')]]"
    ).first

    if row.count() > 0 and row.is_visible():
        print(f"✅ Shipment present in NPR Action Requested page with the shipment AWB number: {awb_text}")
        row_text = row.inner_text()
        assert awb_text in row_text
    else:
        print(f"❌ Shipment not present in NPR Action Requested page: {awb_text}")
        assert False, f"Shipment not found in NPR Action Requested page: {awb_text}"


from utils.config import (
    BASE_URL,
    UAT_URL,
    PROD_URL,
    UAT_CARRIER_URL,
    PROD_CARRIER_URL,
    UAT_CARRIER_USERNAME,
    UAT_CARRIER_PASSWORD,
    PROD_CARRIER_USERNAME,
    PROD_CARRIER_PASSWORD
)


def get_carrier_config():
    if BASE_URL == UAT_URL:
        return UAT_CARRIER_URL, UAT_CARRIER_USERNAME, UAT_CARRIER_PASSWORD

    elif BASE_URL == PROD_URL:
        return PROD_CARRIER_URL, PROD_CARRIER_USERNAME, PROD_CARRIER_PASSWORD

    else:
        raise Exception(f"Unknown BASE_URL: {BASE_URL}")


def update_carrier_tracking_status(page, awb_text):
    # carrier_url, username, password = get_carrier_config()
    # carrier_page = page.context.new_page()
    #
    #
    #
    # carrier_page.goto(carrier_url)
    # carrier_page.wait_for_load_state("networkidle")
    #
    # # Login
    # carrier_page.get_by_role("textbox", name="Enter username").fill(username)
    # carrier_page.get_by_role("textbox", name="Enter Password").fill(password)
    # carrier_page.get_by_role("button", name="Login").click()
    #
    # carrier_page.wait_for_load_state("networkidle")
    #
    # # Go to shipments
    #
    # carrier_page.get_by_role("link", name="local_shipping Shipments").click()
    #
    # # Search AWB
    # search_box = carrier_page.get_by_placeholder("Search By - Tracking No./Trip")
    # search_box.wait_for(state="visible", timeout=10000)
    # search_box.fill(awb_text)
    # carrier_page.keyboard.press("Enter")
    #
    # carrier_page.wait_for_timeout(3000)
    #
    # # Click edit button dynamically
    # carrier_page.get_by_role("cell", name=f"{awb_text} edit_note").locator("#modalbtn").click()
    # time.sleep(3)
    #
    # # Update status
    # carrier_page.locator("#checkpoint_statusis").select_option("PickedUp")
    #
    # carrier_page.wait_for_function(
    #     "() => document.querySelector('#checkpoint_statusis')?.value === 'PickedUp'"
    # )
    #
    #
    # # Scroll modal
    # carrier_page.locator(".modal-body").first.evaluate(
    #     "(el) => el.scrollTop = el.scrollHeight"
    # )
    #
    # # Click update
    # carrier_page.locator("#submitupdate").click()
    # time.sleep(8)
    #
    # print(f"✅ Carrier status updated as PickedUp for AWB: {awb_text}")
    # page.goto("https://uat-carriers.eshipz.com/logout")
    #
    #
    # carrier_page.close()
    # page.bring_to_front()
    # page.goto(f"{BASE_URL}/v2/exceptions/npr")
    page.get_by_text("ExceptionUndelivered").click()
    page.get_by_role("button", name="Tracking").click()
    page.get_by_role("button", name="Update Tracking").click()
    page.get_by_role("combobox", name="Tracking Status *").click()
    page.get_by_role("option", name="Delivered").click()
    page.get_by_role("combobox", name="Tracking Sub Status *").click()
    page.get_by_role("option", name="Delivered").click()
    page.get_by_role("button", name="Update", exact=True).click()

    time.sleep(2)

    page.get_by_role("button", name="Delivered").click()
    page.get_by_role("button", name="Bulk Search").click()
    page.get_by_role("textbox", name="Search").click()
    page.get_by_role("textbox", name="Search").fill(awb_text)
    page.get_by_role("button", name="Search exception").click()
    row = page.locator(
        f"//div[contains(@class,'flex') and .//text()[contains(.,'{awb_text}')]]"
    ).first

    if row.count() > 0 and row.is_visible():
        print(
            f"✅ Shipment has moved from NDR Action Requested page to RTO tab with the shipment AWB number: {awb_text}")
        row_text = row.inner_text()
        assert awb_text in row_text
    else:
        print(f"❌ Shipment not present in NDR Action Requested page : {awb_text}")
        assert False, f"Shipment not found in NDR Action Requested page: {awb_text}"


def update_carrier_tracking_status_forRTO(page, awb_text):
    # carrier_url, username, password = get_carrier_config()
    # carrier_page = page.context.new_page()
    #
    #
    #
    # carrier_page.goto(carrier_url)
    # carrier_page.wait_for_load_state("networkidle")
    #
    # # Login
    # carrier_page.get_by_role("textbox", name="Enter username").fill(username)
    # carrier_page.get_by_role("textbox", name="Enter Password").fill(password)
    # carrier_page.get_by_role("button", name="Login").click()
    #
    # carrier_page.wait_for_load_state("networkidle")
    #
    # # Go to shipments
    #
    # carrier_page.get_by_role("link", name="local_shipping Shipments").click()
    #
    # # Search AWB
    # search_box = carrier_page.get_by_placeholder("Search By - Tracking No./Trip")
    # search_box.wait_for(state="visible", timeout=10000)
    # search_box.fill(awb_text)
    # carrier_page.keyboard.press("Enter")
    #
    # carrier_page.wait_for_timeout(3000)
    #
    # # Click edit button dynamically
    # carrier_page.get_by_role("cell", name=f"{awb_text} edit_note").locator("#modalbtn").click()
    # time.sleep(3)
    #
    # # Update status
    # carrier_page.locator("#checkpoint_statusis").select_option("PickedUp")
    #
    # carrier_page.wait_for_function(
    #     "() => document.querySelector('#checkpoint_statusis')?.value === 'PickedUp'"
    # )
    #
    #
    # # Scroll modal
    # carrier_page.locator(".modal-body").first.evaluate(
    #     "(el) => el.scrollTop = el.scrollHeight"
    # )
    #
    # # Click update
    # carrier_page.locator("#submitupdate").click()
    # time.sleep(8)
    #
    # print(f"✅ Carrier status updated as PickedUp for AWB: {awb_text}")
    # page.goto("https://uat-carriers.eshipz.com/logout")
    #
    #
    # carrier_page.close()
    # page.bring_to_front()
    # page.goto(f"{BASE_URL}/v2/exceptions/npr")
    page.get_by_text("ExceptionUndelivered").click()
    page.get_by_role("button", name="Tracking").click()
    page.get_by_role("button", name="Update Tracking").click()
    page.get_by_role("combobox", name="Tracking Status *").click()
    page.get_by_role("option", name="Return").click()
    page.get_by_role("combobox", name="Tracking Sub Status *").click()
    page.get_by_role("option", name="RTODelivered").click()
    page.get_by_role("button", name="Update", exact=True).click()

    time.sleep(2)

    page.get_by_role("button", name="RTO").click()
    page.get_by_role("button", name="Bulk Search").click()
    page.get_by_role("textbox", name="Search").click()
    page.get_by_role("textbox", name="Search").fill(awb_text)
    page.get_by_role("button", name="Search exception").click()
    row = page.locator(
        f"//div[contains(@class,'flex') and .//text()[contains(.,'{awb_text}')]]"
    ).first

    if row.count() > 0 and row.is_visible():
        print(
            f"✅ Shipment has moved from NDR Action Requested page to RTO tab with the shipment AWB number: {awb_text}")
        row_text = row.inner_text()
        assert awb_text in row_text
    else:
        print(f"❌ Shipment not present in NDR Action Requested page : {awb_text}")
        assert False, f"Shipment not found in NDR Action Requested page: {awb_text}"



from datetime import datetime, timedelta
import random
from datetime import datetime, timedelta
from playwright.sync_api import expect

from datetime import datetime, timedelta
from playwright.sync_api import expect


def fill_npr_reschedule_bulkdate(page):
    print("➡ Filling Reschedule NPR details (Yesterday date)")

    # 🔥 Yesterday date
    date_value = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%dT%H:%M")

    date = page.locator("#reschedule-date")
    expect(date).to_be_visible(timeout=10000)
    expect(date).to_be_enabled(timeout=10000)

    date.fill(date_value)
    print("✅ Reschedule date filled:", date_value)

    comments = page.locator("textarea").first
    expect(comments).to_be_visible(timeout=10000)
    comments.fill("Bulk NPR reschedule test - past date")

    print("✅ Comments filled")


def fill_npr_bulkedit_and_reschedule(page):
    page.get_by_role("button", name="Edit & Re-Schedule").click()
    page.get_by_role("textbox", name="Enter seller name").click()
    page.get_by_role("textbox", name="Enter seller name").fill("cezachian")
    page.get_by_placeholder("Enter phone number").click()
    page.get_by_placeholder("Enter phone number").fill("9898989898")
    page.get_by_role("textbox", name="Enter seller address").click()
    page.get_by_role("textbox", name="Enter seller address").fill("czech street")
    page.get_by_placeholder("Enter pincode").click()
    page.get_by_placeholder("Enter pincode").fill("110002")

    ddate_value = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%dT%H:%M")

    date = page.locator("#edit-reschedule-date")

    expect(date).to_be_visible(timeout=10000)
    expect(date).to_be_enabled(timeout=10000)

    date.fill(ddate_value)
    print("✅ Reschedule date filled:", ddate_value)

    comments = page.locator("textarea").first
    expect(comments).to_be_visible(timeout=10000)
    comments.fill("Bulk NPR reschedule test - past date")


from playwright.sync_api import TimeoutError, expect
import time


def confirm_bulknpr_action(page, awb_text=None):
    print("➡ confirm_bulknpr_action called")

    confirm_btn = page.get_by_role("button", name="Confirm")
    confirm_btn.click()

    page.wait_for_timeout(1000)

    popup = page.locator(
        "//div[@role='status']//span[contains(text(),'scheduled for pickup')]"
    ).last

    try:
        popup.wait_for(state="visible", timeout=5000)

        popup_text = popup.inner_text().strip()
        print("Popup Message:", popup_text)

        assert "Shipment has been scheduled for pickup" in popup_text

        print("✅ Success popup validated")
        print("✅ NPR ACTION Success")


    except TimeoutError:
        print("⚠️ Popup not displayed → going to Action Required page")

    except Exception as e:
        print(f"⚠️ Confirm/action failed: {e}")

    print("➡ Going to Action Required page")

    page.goto("https://uat.eshipz.com:444/v2/exceptions/npr/actionRequired")
    page.wait_for_load_state("networkidle")

    # ✅ perform search using AWB
    if awb_text:
        page.get_by_role("button", name="Bulk Search").click()

        search_box = page.get_by_role("textbox", name="Search")
        search_box.wait_for(state="visible")

        search_box.fill(awb_text)

        page.get_by_role("button", name="Search exception").click()

        print(f"🔍 Searched AWB in Action Required: {awb_text}")
        return "ACTION_REQUIRED"

    time.sleep(3)


from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, timedelta
import random


def download_bulk_npr_excel(page):
    print("➡ Opening Bulk NPR upload")

    download_dir = Path("downloadedfiles")
    download_dir.mkdir(exist_ok=True)

    page.get_by_role("button", name="Bulk upload").click()

    with page.expect_download() as download_info:
        page.get_by_role("button", name="NPR Download excel_download").click()

        # from datetime import datetime
        #
        # today = str(datetime.now().day)
        # date_btn = page.get_by_role("button", name=today).nth(2)
        #
        # date_btn.click()
        # date_btn.click()
        page.locator("//label[normalize-space()='Today']").click()

        page.get_by_role("button", name="Apply").click()

    download = download_info.value
    file_path = download_dir / download.suggested_filename
    download.save_as(file_path)

    print(f"✅ NPR Excel downloaded: {file_path}")
    return str(file_path)


from openpyxl import load_workbook
from datetime import datetime, timedelta
import random


def update_bulk_npr_excel(excel_path, action_type):
    print(f"➡ Updating Bulk NPR Excel with action: {action_type}")

    wb = load_workbook(excel_path)
    ws = wb.active

    headers = {}

    # headers are in row 1
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip()] = col

    action_col = headers.get("Action")
    phone_col = headers.get("Phone") or headers.get("Phone Number")
    name_col = headers.get("Name")
    address_col = headers.get("Address")
    pincode_col = headers.get("Pincode") or headers.get("New Pincode")
    pickup_date_col = headers.get("PickupDateTime")

    if not action_col:
        raise Exception("Action column not found")

    if not pickup_date_col:
        raise Exception("PickupDateTime column not found")

    if action_type == "RESCHEDULE":
        excel_action = "RE-SCHEDULE"

    elif action_type == "EDIT_AND_RESCHEDULE":
        excel_action = "EDIT_DETAILS"

    elif action_type == "CANCEL":
        excel_action = "CANCEL"

    else:
        raise Exception(f"Invalid action type: {action_type}")

    # ✅ start from 3rd row, because row 2 has instruction text
    for row in range(3, ws.max_row + 1):

        # skip empty shipment rows
        message_value = ws.cell(row=row, column=1).value
        if not message_value:
            continue

        ws.cell(row=row, column=action_col).value = excel_action

        # ✅ tomorrow or day after tomorrow
        future_date = datetime.now() + timedelta(days=random.choice([1, 2]))

        # ✅ random morning time between 10:00 AM and 11:59 AM
        hour = random.choice([10, 11])
        minute = random.randint(0, 59)

        # ✅ Excel required format: YYYY/MM/DD HH:MM
        pickup_date = future_date.replace(
            hour=hour,
            minute=minute
        ).strftime("%Y/%m/%d %H:%M")

        if excel_action == "RE-SCHEDULE":
            ws.cell(row=row, column=pickup_date_col).value = pickup_date

        elif excel_action == "EDIT_DETAILS":
            random_phone = "9" + "".join(str(random.randint(0, 9)) for _ in range(9))

            if phone_col:
                ws.cell(row=row, column=phone_col).value = random_phone

            if name_col:
                ws.cell(row=row, column=name_col).value = "Raj Kumar"

            if address_col:
                ws.cell(row=row, column=address_col).value = "Bangalore Test Address"

            if pincode_col:
                ws.cell(row=row, column=pincode_col).value = "110002"

            ws.cell(row=row, column=pickup_date_col).value = pickup_date

        elif excel_action == "CANCEL":
            pass

    wb.save(excel_path)

    print(f"✅ Excel updated successfully: {excel_path}")
    print(f"✅ Action filled from row 3: {excel_action}")

    return excel_path


from openpyxl import load_workbook
from pathlib import Path


def get_waybills_from_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip()] = col

    waybill_col = headers.get("Waybill")
    if not waybill_col:
        raise Exception("Waybill column not found in Excel")

    waybills = []

    # start from row 3 because row 2 is DON'T CHANGE
    for row in range(3, ws.max_row + 1):
        value = ws.cell(row=row, column=waybill_col).value

        if value:
            waybills.append(str(value).strip())

    waybill_text = ",".join(waybills)

    print("✅ Waybills collected from Excel:")
    print(waybill_text)

    return waybill_text


def upload_bulk_npr_excel(page, excel_path):
    print("➡ Uploading updated Bulk NPR Excel")

    # page.get_by_text("Browse File").click()

    # page.locator("input[type='file']").set_input_files(str(Path(excel_path).resolve()))
    file_input = page.locator("input[type='file']").first
    file_input.set_input_files(str(Path(excel_path).resolve()))

    page.wait_for_timeout(3000)

    print("✅ Bulk NPR Excel uploaded successfully")


def search_waybills_in_action_requested(page, waybill_text):
    print("➡ Searching waybills in Action Requested")

    # ✅ Check in Action Requested (SUCCESS case)
    page.get_by_role("button", name="Action Requested").click()
    page.wait_for_timeout(2000)

    page.get_by_role("button", name="Bulk Search").click()

    search_box = page.get_by_role("textbox", name="Search")
    search_box.wait_for(state="visible", timeout=10000)
    search_box.fill(waybill_text)

    page.get_by_role("button", name="Search exception").click()
    page.wait_for_timeout(3000)

    rows = page.locator("table tbody tr")

    if rows.count() > 0:
        print("✅ SUCCESS: Waybills found in Action Requested")
        return "SUCCESS"

    # ❌ If not found → check Action Required
    print("⚠️ Not found in Action Requested → checking Action Required")

    page.get_by_role("button", name="Action Required").click()
    page.wait_for_timeout(2000)

    page.get_by_role("button", name="Bulk Search").click()

    search_box = page.get_by_role("textbox", name="Search")
    search_box.wait_for(state="visible", timeout=10000)
    search_box.fill(waybill_text)

    page.get_by_role("button", name="Search exception").click()
    page.wait_for_timeout(3000)

    rows = page.locator("table tbody tr")

    if rows.count() > 0:
        print("❌ FAILURE: Waybills found in Action Required")
        return "FAILED"

    print("⚠️ No records found in both tabs")
    return "NOT_FOUND"